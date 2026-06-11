# ============================================================
# main.py — boAt Crest App Review Analyzer (Cloud Version)
# Author  : boAt R&D (QA)
# Replaces: CrestApp_WebScraping, All_HtmlFiles_Csv,
#           App_Rating_Finder, Csv_Date_Rating_Filter_7days,
#           CSV_Remove_Duplicates, Csv_2_Xlsx
# ============================================================

import os
import uuid
import time
import pytz
import openpyxl
import pandas as pd
import BFAT_CrestApp_PlayStore_Code as orig
from datetime import datetime, timedelta
from google_play_scraper import Sort, reviews

from BFAT_CrestApp_PlayStore_Code import (
    SheetNameChange,
    Positive_Reviews,
    Negative_Reviews,
    Sorting_Xlsx,
    Sorting_Summary_Xlsx,
    Negative_Review_Analysis,
    Review_Date_Calculator,
    Summary,
    Summary_Enhance,
    close_excel_file,
)

# ── CONFIG ──────────────────────────────────────────────────
APP_ID    = 'com.coveiot.android.boat'
XLSX_FILE = 'Android_CrestApp_Review.xlsx'

# ── GLOBALS ─────────────────────────────────────────────────
App_Rating  = "N/A"
App_Version = "N/A"
lowest_date = None
latest_date = None

# ── STEP 1: FETCH REVIEWS ───────────────────────────────────
def fetch_and_save_reviews():
    print("=" * 60)
    print("Step 1: Fetching reviews from Google Play Store...")
    print("=" * 60)

    result, _ = reviews(
        APP_ID,
        lang='en',
        country='in',
        sort=Sort.NEWEST,
        count=600,
    )

    df = pd.DataFrame(result)

    # ── FIX: Use IST timezone so GitHub Actions (UTC) and local machine
    #         both use the same cutoff. Without this, GitHub's clock is
    #         5.5 hours behind IST, causing recent reviews to be missed.
    ist = pytz.timezone('Asia/Kolkata')
    now_ist = datetime.now(ist).replace(tzinfo=None)
    seven_days_ago = now_ist - timedelta(days=7)

    # Play Store API returns timezone-aware datetimes — strip tzinfo
    # so both sides are plain datetime for comparison
    df['at'] = pd.to_datetime(df['at']).dt.tz_localize(None)
    df = df[df['at'] >= seven_days_ago]

    print(f"  IST now        : {now_ist.strftime('%Y-%m-%d %H:%M:%S')} IST")
    print(f"  Cutoff (7 days): {seven_days_ago.strftime('%Y-%m-%d %H:%M:%S')} IST")

    # Add unique reviewId
    df['reviewId'] = [str(uuid.uuid4()) for _ in range(len(df))]

    # Rename columns
    df = df.rename(columns={
        'userName':             'User Name',
        'score':                'Ratings',
        'content':              'Reviews',
        'at':                   'Date Time',
        'reviewCreatedVersion': 'App Version',
        'replyContent':         'boAt Reply',
    })

    # Keep only useful columns
    df = df[[
        'reviewId', 'User Name', 'Ratings', 'Reviews',
        'Date Time', 'App Version', 'boAt Reply'
    ]]

    # Sort and deduplicate
    df = df.sort_values('Date Time', ascending=False)
    df = df.drop_duplicates(subset=['User Name', 'Ratings', 'Reviews'])

    # Save to Excel
    df.to_excel(XLSX_FILE, index=False)

    print(f"  Fetched : {len(df)} reviews from last 7 days")
    print(f"  Saved   : {XLSX_FILE}")

# ── STEP 2: GET APP VERSION ─────────────────────────────────
def App_Version_Finder():
    global App_Version

    workbook = openpyxl.load_workbook(XLSX_FILE)
    sheet    = workbook.active

    APP_VERSION_COL = 6   # column F — App Version
    highest_version = None

    for row in range(2, sheet.max_row + 1):
        ver = sheet.cell(row=row, column=APP_VERSION_COL).value
        if ver:
            ver_str = str(ver).split(":")[-1].strip()
            if highest_version is None or ver_str > highest_version:
                highest_version = ver_str

    App_Version = highest_version if highest_version else "N/A"
    print(f"  Highest App Version: {App_Version}")
    workbook.close()

# ── MAIN ────────────────────────────────────────────────────
if __name__ == '__main__':

    # Step 1 — Fetch reviews from Play Store
    fetch_and_save_reviews()
    time.sleep(2)

    # Step 2 — Rename Sheet1 → CrestApp_Review_Data
    SheetNameChange()

    # Step 3 — Split into Positive / Negative sheets
    Positive_Reviews()
    Negative_Reviews()

    # Step 4 — Format all sheets
    Sorting_Xlsx('Positive_Reviews')
    Sorting_Xlsx('Negative_Reviews')
    Sorting_Xlsx('CrestApp_Review_Data')

    # Step 5 — Analysis
    Negative_Review_Analysis()

    # Step 6 — Get App Version + Date Range
    App_Version_Finder()
    Review_Date_Calculator()

    # Step 7 — Inject globals into original script before Summary
    orig.App_Version = App_Version
    orig.App_Rating  = App_Rating
    orig.lowest_date = lowest_date
    orig.latest_date = latest_date

    # Step 8 — Generate Summary sheet
    Summary()
    time.sleep(3)

    # Step 9 — Format and enhance Summary
    Sorting_Summary_Xlsx('Summary')
    time.sleep(3)
    Summary_Enhance()
    time.sleep(3)

    # Step 10 — Close file
    close_excel_file(XLSX_FILE)

    print("=" * 60)
    print("Pipeline complete! Android_CrestApp_Review.xlsx ready.")
    print("=" * 60)
