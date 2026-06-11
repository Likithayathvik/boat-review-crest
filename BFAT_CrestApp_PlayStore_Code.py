#Project Name : boAt - App Feedback Analyzer Tool
#Author       : boAt R&D (QA)
#Date         : 26 Nov 2024
#Version      : v2.3
#Description  : Google Console Play Store boAt App Feedback Analyzer
#Script & IDE : Python - Web Scraping [IDE Used - Thonny]
#********************************************************************#

#Library dependencies
#---------------------
import os
import re
import csv
import sys
import time
import uuid
import shutil
import smtplib
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from selenium.webdriver.common.by import By
from email.mime.multipart import MIMEMultipart
from selenium.webdriver.common.keys import Keys
from email.mime.application import MIMEApplication
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl.styles import Font, PatternFill, Alignment
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import ElementClickInterceptedException
import time
import undetected_chromedriver as uc

# Global variables declaration
global App_Rating
global App_Version
global lowest_date
global latest_date

#***********************************************************************#

def CrestApp_WebScraping():
    # Create a Chrome instance using undetected-chromedriver
    options = uc.ChromeOptions()
    options.add_argument("--no-first-run --no-service-autorun --password-store=basic")

    # Start the driver
    driver = uc.Chrome(options=options, version_main=147)

    try:
        # Navigate directly to Play Console — Google will redirect to sign-in if needed
        driver.get('https://play.google.com/console/u/0/developers/6824393389065942833/app/4972589525526478206/user-feedback/reviews')
        time.sleep(10)

        # Find the email input field and enter your email
        email_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'identifierId'))
        )
        email_input.send_keys('boatqatest@gmail.com')
        email_input.send_keys(Keys.RETURN)

        time.sleep(5)

        # Wait for the password input field and enter password
        password_input = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.NAME, 'Passwd'))
        )
        password_input.send_keys('Testing@12345')
        password_input.send_keys(Keys.RETURN)

        time.sleep(10)

        try:
            continue_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='Continue']"))
            )
            continue_button.click()
            print("Clicked Continue")
            time.sleep(5)
        except:
            print("No Continue button found or popup didn't appear")

        # Handle pop-up and click "Cancel"
        try:
            cancel_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='Cancel']"))
            )
            cancel_button.click()
            print("Cancelled popup")
            time.sleep(5)
        except:
            print("No Cancel button found or popup didn't appear")

        # Now, navigate to the desired page (in this case, Google Play Console)
        driver.get('https://play.google.com/console/u/0/developers/6824393389065942833/app/4972589525526478206/user-feedback/reviews')
        time.sleep(10)

        # Wait for the dropdown button to be clickable
        dropdown_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//div[@aria-label="Reviews per page 10 selected."]'))
        )

        # Click the dropdown button
        dropdown_button.click()
        print("Dropdown clicked")
        time.sleep(5)

        # IMPROVED: Use a more reliable selector for the "50" option
        try:
            # Wait for the dropdown options to appear
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'material-select-dropdown-item'))
            )

            # Find the "50" option by its text content
            option_50 = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//material-select-dropdown-item[.//span[text()='50']]"))
            )

            # Scroll into view if needed
            driver.execute_script("arguments[0].scrollIntoView(true);", option_50)
            time.sleep(1)

            # Try clicking
            try:
                option_50.click()
                print("Successfully clicked 50 items per page option")
            except ElementClickInterceptedException:
                # If regular click fails, use JavaScript click
                driver.execute_script("arguments[0].click();", option_50)
                print("Successfully clicked 50 items per page option using JavaScript")

            time.sleep(10)

            # Save the webpage to an HTML file
            with open("CrestApp_1.html", "w", encoding="utf-8") as file:
                file.write(driver.page_source)
            print("Web page contents saved to 'CrestApp_1.html'.")

            # Get the HTML content of the page after navigating to it
            html_content = driver.page_source

            # Parse the HTML using BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')

            # Define the CSS selector for the material-button with the specified aria-label
            css_selector = 'material-button[aria-label="Go to the next page"]'

            # Loop to click the element 14 times
            for i in range(14): #if user wants to load the review page, here can edit [Ex: 3pages means change the range to 3]
                # Wait for the element to be present on the page
                wait = WebDriverWait(driver, 10)
                element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, css_selector)))

                if element:
                    try:
                        # Check if the element is clickable
                        if element.is_enabled() and element.is_displayed():
                            # Click the element
                            element.click()
                            print(f"Clicked the material-button with aria-label 'Go to the next page' (Click {i + 1}/14).")
                            time.sleep(10)
                            # Save the page source to an HTML file
                            with open(f'CrestApp_{i+2}.html', 'w', encoding='utf-8') as file:
                                file.write(driver.page_source)
                            print(f"Page details saved in 'CrestApp_{i+2}.html'.")
                        else:
                            print("The material-button is not clickable.")
                    except ElementClickInterceptedException as e:
                        print("Element click intercepted: ", e)
                else:
                    print("The material-button with aria-label 'Go to the next page' is not available on the page.")

        except Exception as e:
            print(f"Error selecting 50 items per page: {e}")
            print("Trying alternative method...")

            # Alternative method - find all items and click the one with text "50"
            try:
                all_items = driver.find_elements(By.CSS_SELECTOR, 'material-select-dropdown-item span.label')

                for item in all_items:
                    if item.text.strip() == '50':
                        try:
                            item.click()
                            print("Clicked on 50 using alternative method")
                            time.sleep(10)
                            break
                        except:
                            driver.execute_script("arguments[0].click();", item)
                            print("Clicked on 50 using JavaScript (alternative method)")
                            time.sleep(10)
                            break
            except Exception as e2:
                print(f"Alternative method also failed: {e2}")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close the WebDriver
        driver.quit()

def find_elements_in_review_container(html_content):

    global App_Rating_Value

    soup = BeautifulSoup(html_content, 'html.parser')
    review_containers = soup.find_all(class_='review-container')

    data = []

    for container in review_containers:
        App_Rating = container.find(attrs={'debug-id': 'google-play-rating'})
        App_Rating_Value = App_Rating.find('span', class_='value').text.strip() if App_Rating else "NA"

        author_display_name = container.find(class_='author-display-name')
        last_update_time = container.find(class_='last-update-time')
        device_display_name = container.find(attrs={'debug-id': 'device-display-name'})
        version_code = container.find(attrs={'debug-id': 'version-code'})
        version_name = container.find(attrs={'debug-id': 'version-name'})
        os_version = container.find(attrs={'debug-id': 'os-version'})
        review_body = container.find(attrs={'debug-id': 'review-body'})

        star_icons = container.find(attrs={'debug-id': 'star-icons'})
        star_icons_value = star_icons.get('aria-label').split(' ')[0] if star_icons else "NA"

        conversation_content = container.find(class_='conversation-content')

        # If any element is not found, set its value to "NA"
        author_display_name_value = author_display_name.text.strip() if author_display_name else "NA"
        last_update_time_value = last_update_time.text.strip() if last_update_time else "NA"
        device_display_name_value = device_display_name.text.strip() if device_display_name else "NA"
        version_code_value = version_code.text.strip() if version_code else "NA"
        version_name_value = version_name.text.strip() if version_name else "NA"
        os_version_value = os_version.text.strip() if os_version else "NA"
        review_body_value = review_body.text.strip() if review_body else "NA"
        conversation_content_value = conversation_content.text.strip() if conversation_content else "NA"

        # Append data to the list
        data.append([
            author_display_name_value,
            star_icons_value,
            review_body_value,
            device_display_name_value,
            os_version_value,
            last_update_time_value,
            version_code_value,
            version_name_value,
            conversation_content_value
        ])

    return data

def All_HtmlFiles_Csv():
    directory_path = os.path.dirname(os.path.realpath(__file__))
    folder_path = directory_path  # Replace with the path to your folder

    # Get a list of all files in the folder
    file_list = os.listdir(folder_path)

    # Filter the list to include only HTML files
    html_files = [file for file in file_list if file.endswith('.html')]

    # List to store data from all HTML files
    all_data = []

    # Process each HTML file
    for html_file in html_files:
        # Wait for 10 seconds before processing the next file
        time.sleep(10)
        file_path = os.path.join(folder_path, html_file)
        with open(file_path, 'r', encoding='utf-8') as file:
            html_content = file.read()
        print(f"Processing: {file_path}")

        # Get data for the current HTML file
        data = find_elements_in_review_container(html_content)

        # Append data to the list for all HTML files
        all_data.extend(data)

    # Write data to a single CSV file
    csv_file_path = os.path.join(directory_path, 'Crest_Data.csv')
    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
        csv_writer = csv.writer(csvfile)

        # Write header
        header = ['User Name', 'Ratings', 'Reviews',
                  'Phone Name', 'OS Version', 'Date Time',
                  'Version Code', 'App Version', 'boAt Reply']
        csv_writer.writerow(header)

        # Write data rows
        csv_writer.writerows(all_data)

def Csv_Date_Rating_Filter_Alldate():
    # Load the CSV file into a DataFrame
    df = pd.read_csv('Crest_Data.csv')

    # Convert the "Date Time" column to a datetime data type if it's not already
    df['Date Time'] = pd.to_datetime(df['Date Time'])
    df['Date Time'] = pd.to_datetime(df['Date Time'], format="%b %d, %Y, %H:%M")

    # Remove the filtering for the last 7 days
    # Comment out or remove the following two lines
    # seven_days_ago = datetime.now() - timedelta(days=7)
    # df = df[df['Date Time'] >= seven_days_ago]

    # Sort the DataFrame in descending order based on the "Date Time" column
    df = df.sort_values(by='Date Time', ascending=False)

    # Filter the "Ratings" column to keep only values 1, 2, and 3
    df = df[df['Ratings'].isin([1, 2, 3, 4, 5])]

    # Save the filtered DataFrame to a new CSV file
    df.to_csv('CrestAppGCWeb.csv', index=False)

    time.sleep(2)

    # Load the CSV file into a pandas DataFrame
    df = pd.read_csv('CrestAppGCWeb.csv')

    # Create a new column 'reviewId' with unique identifiers
    df['reviewId'] = [str(uuid.uuid4()) for _ in range(len(df))]

    # Reorder the columns to place 'reviewId' in the 'A' column
    cols = df.columns.tolist()
    cols = ['reviewId'] + [col for col in cols if col != 'reviewId']
    df = df[cols]

    # Save the DataFrame with 'reviewId' column to the same CSV file
    df.to_csv('CrestAppGCWeb.csv', index=False)

    print("Date and Ratings filter is successfully completed and saved into CrestAppGCWeb CSV file")


def Csv_Date_Rating_Filter_7days():
    # Load the CSV file into a DataFrame
    df = pd.read_csv('Crest_Data.csv')
    # Convert the "Date Time" column to a datetime data type if it's not already
    #df['Date Time'] = pd.to_datetime(df['Date Time'])
    df['Date Time'] = pd.to_datetime(df['Date Time'], format="%b %d, %Y, %H:%M")
    # Calculate the date 7 days ago from the current date
    seven_days_ago = datetime.now() - timedelta(days=7)
    # Filter the DataFrame to keep only rows with dates from the last 7 days
    filtered_df = df[df['Date Time'] >= seven_days_ago]
    # Sort the filtered DataFrame in descending order based on the "Date Time" column
    #filtered_df = filtered_df.sort_values(by='Date Time', ascending=False)
    # Filter the "Ratings" column to keep only values 1 to 5
    filtered_df = filtered_df[filtered_df['Ratings'].isin([1, 2, 3, 4, 5])]
    # Save the filtered DataFrame to a new CSV file
    filtered_df.to_csv('CrestAppGCWeb.csv', index=False)
    time.sleep(2)
    # Load the CSV file into a pandas DataFrame
    df = pd.read_csv('CrestAppGCWeb.csv')
    # Create a new column 'reviewId' with unique identifiers
    df['reviewId'] = [str(uuid.uuid4()) for _ in range(len(df))]
    # Reorder the columns to place 'reviewId' in the 'A' column
    cols = df.columns.tolist()
    cols = ['reviewId'] + [col for col in cols if col != 'reviewId']
    df = df[cols]
    df.to_csv('CrestAppGCWeb.csv', index=False)

    print("Date and Ratings filter is successfully completed and saved into CrestAppGCWeb CSV file")

def CSV_Remove_Duplicates():
    # Get the current directory path
    directory_path = os.path.dirname(os.path.realpath(__file__))

    # Specify the folder path where the CSV files are located
    folder_path = directory_path  # Replace with the path to your folder

    # Check if the folder path exists
    if os.path.exists(folder_path):
        # List all files in the folder
        files = os.listdir(folder_path)

        # Loop through each file in the folder
        for file_name in files:
            # Check if the file is a CSV file
            if file_name.endswith('.csv'):
                # Construct the full path to the CSV file
                file_path = os.path.join(folder_path, file_name)

                # Read the CSV file into a pandas DataFrame
                df = pd.read_csv(file_path)

                # Count and remove duplicates based on the specified columns
                duplicate_count_before = df.duplicated(subset=['User Name', 'Ratings', 'Reviews', 'Phone Name']).sum()
                df = df.drop_duplicates(subset=['User Name', 'Ratings', 'Reviews', 'Phone Name'])

                # Overwrite the original CSV file with the cleaned DataFrame
                df.to_csv(file_path, index=False)

                # Print the number of duplicate rows removed for each file
                print(f'For {file_name}: {duplicate_count_before} duplicate rows removed.')
    else:
        print(f'The specified folder path "{folder_path}" does not exist.')

def Csv_2_Xlsx():

    directory_path = os.path.dirname(os.path.realpath(__file__))  # Get the current script's directory
    csv_file = 'CrestAppGCWeb.csv'
    xlsx_file = 'Android_CrestApp_Review.xlsx'
    # Read the CSV file using pandas.
    df = pd.read_csv(csv_file)

    # Write the DataFrame to an XLSX file.
    df.to_excel(xlsx_file, index=False)

    print(f'CSV file "{csv_file}" has been converted to XLSX file "{xlsx_file}".')

def Positive_Reviews():
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_CrestApp_Review.xlsx')

    # Filter the DataFrame to keep only rows with ratings 4 and 5
    positive_reviews_df = df[df['Ratings'].isin([4, 5])]

    # Save the positive reviews to the same Excel file in a new sheet named 'Positive_Reviews'
    with pd.ExcelWriter('Android_CrestApp_Review.xlsx', engine='openpyxl', mode='a') as writer:
        positive_reviews_df.to_excel(writer, sheet_name='Positive_Reviews', index=False)

    print("Positive reviews saved to 'Android_CrestApp_Review.xlsx' under sheet name 'Positive_Reviews'.")

def Negative_Reviews():
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_CrestApp_Review.xlsx')

    # Filter the DataFrame to keep only rows with ratings 1, 2, and 3
    negative_reviews_df = df[df['Ratings'].isin([1, 2, 3])].copy()  # Make a copy to avoid modifying the original DataFrame

    # Keyword patterns
    keyword_patterns = {
    'Connectivity': r'\b(connect(ion|ing)?|not\s*connected|not\s*getting\s*connect(ed|ion)?|unable\s*to\s*connect|bad\s*connectivity)\b',
    'Device disconnection': r'\bdisconnected?\b',
    'Login': r'\b(login|log(in|ged)?|otp|sign\s*in|signing\s*in)\b',
    'Location Permission': r'\b(location|gps|geolocation)\s*(permission|access)?\b',
    'Phone becoming silent': r'\bsilent\b',
    'SKU Not supported': r"\b(can't\s*connect.*app|not\s*(connect(ing)?|support(ed)?|show(n|ing)|found|recognized|compatible|listed|list|option)\s*(by\s*this\s*app)?|\bthis\s*app\s*doesn't\s*support\s*all\s*boat\s*smartwatch|\bwatch\s*doesn't\s*pair\s*and\s*show\s*in\s*the\s*app|\bboat\s*watch\s*not\s*listed|\bdevices\s*could\s*not\s*found|\bnot\s*available\s*in\s*list|\bsupported\s*device\s*list\b|\bthis\s*model\s*is\s*not\s*(listed|shown|supported|found)\s*in\s*the\s*app\b|\bthis\s*model\s*does\s*not\s*show\s*this\s*model\s*on\s*app\b|\bcannot\s*pair\s*it\s*with\s*this\s*app\b|\bnot\s*listed\s*in\s*app\b|\bnot\s*list\b|\bno\s*option\b|\boption\s*is\s*not\s*available\s*in\s*this\s*app\b|can't\s*see\b|\bdidn'?t\s*find\s*my\s*device\b|\bdoes\s*not\s*show\s*other\s*models\b|\bunable\s*to\s*find\s*my\s*product\b|\bdidn'?t\s*found\s*my\s*device\b|\bcan't\s*found\s*my\s*device\b|\bno\s*support\s*for\s*old\s*devices\b|\bnot\s*supported\s*for\s*old\s*devices\b)",
    'Positive Feedback': r'^(?!.*(not|never|no)).*\b(good|super|nice|nyc|best|beautiful|great|amazing|awesome|impressed)\b',
    'Powering off (Device)': r'\b(auto\s*on\s*off|powering\s*off|auto\s*shutdown)\b',
    'Crash': r'\b(crash(ing)?|hang(ing)?|freeze(ing)?|carshed)\b',
    'Package box (Device)': r'\b(open\s*box|package\s*box)\b',
    'Battery': r'\b(battery|battery\s*percentage|battery\s*level|case\s*battery|watch\s*battery)\b',
    'Watch Face': r'\b(watch\s*faces?|cloud\s*watch\s*face|clock\s*face|smartwatch\s*faces?|watch\s*display|custom\s*watch\s*faces?)\b',
    'Firmware': r'\b(firmware|firm\s*ware)\b'
}


    # Compile patterns with re.IGNORECASE flag
    compiled_patterns = {keyword: re.compile(pattern, re.IGNORECASE) for keyword, pattern in keyword_patterns.items()}

    # Analyze reviews and boAt replies
    reasons = []
    final_reasons = []
    boat_reply_reasons = []  # To store boAt Reply reasons

    # Loop through the reviews and boAt replies
    for review, reply in zip(negative_reviews_df['Reviews'], negative_reviews_df['boAt Reply']):
        found_reasons = []
        found_reply_reasons = []

        # Ensure reply is a string (replace NaN or None with empty string)
        if pd.isna(reply):
            reply = ''

        # Check patterns in the reviews
        for keyword, pattern in keyword_patterns.items():
            if re.search(pattern, review, re.IGNORECASE):
                found_reasons.append(keyword)

        reasons.append(', '.join(found_reasons) if found_reasons else 'Generic Comments')

        # Check patterns in the boAt replies
        for keyword, pattern in keyword_patterns.items():
            if re.search(pattern, reply, re.IGNORECASE):
                found_reply_reasons.append(keyword)

        boat_reply_reasons.append(', '.join(found_reply_reasons) if found_reply_reasons else 'Generic Reply')

        # Final reason logic:
        # If the boAt reply contains "SKU Not supported", override the final reason with "SKU Not supported"
        if 'SKU Not supported' in found_reply_reasons:
            final_reasons.append('SKU Not supported')
        elif len(found_reasons) > 1:
            final_reasons.append('Decision yet to be confirmed')
        else:
            final_reasons.append(', '.join(found_reasons) if found_reasons else 'Generic Comments')

    # Add the columns to the DataFrame
    negative_reviews_df['Reason'] = reasons
    #negative_reviews_df['boAt Reply Reasons'] = boat_reply_reasons
    negative_reviews_df['Consolidated Reason'] = final_reasons

    # Save the negative reviews with reasons to the same Excel file in a new sheet named 'Negative_Reviews'
    with pd.ExcelWriter('Android_CrestApp_Review.xlsx', engine='openpyxl', mode='a') as writer:
        negative_reviews_df.to_excel(writer, sheet_name='Negative_Reviews', index=False)

    print("Negative reviews and reasons saved to 'Negative_Reviews' sheet.")

def SheetNameChange():
    directory_path = os.path.dirname(os.path.realpath(__file__))  # Get the current script's directory
    # Load the Excel file
    file_path = 'Android_CrestApp_Review.xlsx'
    current_sheet_name = 'Sheet1'
    new_sheet_name = 'CrestApp_Review_Data'

    try:
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Check if the current sheet name exists
        if current_sheet_name in wb.sheetnames:
            # Get the sheet object
            sheet = wb[current_sheet_name]

            # Change the sheet name to 'Negative_Reviews'
            sheet.title = new_sheet_name

            # Save the modified workbook with the new sheet name
            wb.save(file_path)

            print(f"The sheet name '{current_sheet_name}' has been changed to '{new_sheet_name}'.")
        else:
            print(f"The sheet '{current_sheet_name}' does not exist in the Excel file.")

        # Close the Excel file
        wb.close()

    except FileNotFoundError:
        print(f"The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


def Sorting_Xlsx(sheet_name):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_CrestApp_Review.xlsx', sheet_name=sheet_name)

    # Load the XLSX file
    workbook = openpyxl.load_workbook('Android_CrestApp_Review.xlsx')

    # Select the worksheet where you want to set column widths
    worksheet = workbook[sheet_name]

    # Define the column widths you want to set (e.g., columns A, B, and C)
    column_widths = {'A': 15, 'B': 15, 'C': 8, 'D': 42, 'E': 12.5, 'F': 11, 'G': 11, 'H': 10, 'I': 10, 'J': 48, 'K': 15, 'L': 21} # You can adjust the values as needed

    # Set the column widths
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # Iterate through the rows and columns to apply wrap text and middle align
    for row in worksheet.iter_rows():
        for cell in row:
            # Set wrap text to True
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")

    # Save the workbook with changes to the same file
    workbook.save('Android_CrestApp_Review.xlsx')

    # Close the workbook
    workbook.close()

    print(f"Xlsx modification done for sheet '{sheet_name}'.")

def Sorting_Summary_Xlsx(sheet_name):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_CrestApp_Review.xlsx', sheet_name=sheet_name)

    # Load the XLSX file
    workbook = openpyxl.load_workbook('Android_CrestApp_Review.xlsx')

    # Select the worksheet where you want to set column widths
    worksheet = workbook[sheet_name]

    # Define the column widths you want to set (e.g., columns A, B, and C)
    column_widths = {'A': 25, 'B': 15, 'C': 21, 'D': 23, 'E': 25, 'F': 28} # You can adjust the values as needed

    # Set the column widths
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # Iterate through the rows and columns to apply wrap text and middle align
    for row in worksheet.iter_rows():
        for cell in row:
            # Set wrap text to True
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")

    # Save the workbook with changes to the same file
    workbook.save('Android_CrestApp_Review.xlsx')

    # Close the workbook
    workbook.close()

    print(f"Xlsx modification done for sheet '{sheet_name}'.")


def Remove_Html():

    directory_path = os.path.dirname(os.path.realpath(__file__))

    folder_path = directory_path  # Replace with the path to your folder

    # Check if the folder path exists
    if os.path.exists(folder_path):
        # List all files in the folder
        files = os.listdir(folder_path)

        # Iterate through the files and delete HTML files
        for file in files:
            if file.endswith(".html"):
                file_path = os.path.join(folder_path, file)
                os.remove(file_path)
                print(f"Deleted: {file_path}")
    else:
        print(f"The folder path '{folder_path}' does not exist.")

def AppRating_find_element_in_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all div elements with class "value-and-description"
    div_elements = soup.find_all('div', class_='value-and-description')

    # Check if any span element with class "value" is found within the div elements
    for div_element in div_elements:
        span_elements = div_element.find_all('span', class_='value')
        if span_elements:
            return span_elements[0].get_text()

    return None

def App_Rating_Finder():
    global App_Rating

    # Path to your HTML file
    html_file_path = 'CrestApp_1.html'

    # Check if file exists before trying to open it
    if not os.path.exists(html_file_path):
        print(f"ERROR: {html_file_path} not found!")
        App_Rating = "N/A"
        return

    # Check if the element is available in the HTML file
    found_element = AppRating_find_element_in_html(html_file_path)

    if found_element:
        App_Rating = found_element[:3]
        print("App Rating Value found in the HTML file:", App_Rating)
    else:
        print("App Rating Value not found in HTML file")
        App_Rating = "N/A"

def App_Version_Finder():
    global App_Version

    # Load the Excel file
    workbook = openpyxl.load_workbook('Android_CrestApp_Review.xlsx')
    sheet = workbook.active

    # Assuming "App Version" is in column I (9)
    app_version_column = 9
    data_start_row = 2

    # Initialize variable to store highest version
    highest_version = None

    # Function to clean and convert version string
    def parse_version(version_str):
        if not version_str:
            return None

        version_str = str(version_str).strip()

        # Handle invalid values
        if version_str in ["-", "N/A", "n/a", "NA", "na"]:
            return None

        # Remove everything except digits and dots (e.g., "4.0-beta" → "4.0")
        version_clean = re.sub(r"[^0-9.]", "", version_str)

        if not version_clean:
            return None

        try:
            return tuple(map(int, version_clean.split('.')))
        except ValueError:
            return None

    # Iterate rows and find highest version
    for row in range(data_start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=app_version_column).value
        parsed_version = parse_version(cell_value)
        if parsed_version:
            if highest_version is None or parsed_version > highest_version:
                highest_version = parsed_version

    # Convert back to string safely
    if highest_version:
        App_Version = '.'.join(map(str, highest_version))
    else:
        App_Version = "Unknown"

    print("Highest App Version:", App_Version)

def Review_Date_Calculator():
    global lowest_date
    global latest_date
    # Load the Excel file
    df = pd.read_excel('Android_CrestApp_Review.xlsx')

    # Replace 'DateTimeColumn' with the actual name of the column containing your date and time data
    df['Date Time'] = pd.to_datetime(df['Date Time'])

    # Extract date part
    df['DateOnly'] = df['Date Time'].dt.date

    # Find the earliest and latest date
    lowest_date = df['DateOnly'].min()
    latest_date = df['DateOnly'].max()

    # Print the results
    print("Lowest Date in the DateTimeColumn:", lowest_date)
    print("Latest Date in the DateTimeColumn:", latest_date)


def Negative_Review_Analysis():
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_CrestApp_Review.xlsx', sheet_name='Negative_Reviews')

    # Count occurrences of each "Consolidated Reason"
    reason_counts = df['Consolidated Reason'].value_counts()

    # Print the counts for each "Consolidated Reason"
    print("Negative Review Analysis:")
    for reason, count in reason_counts.items():
        print(f"{reason}: {count}")

    return reason_counts

def Summary():
    global App_Rating
    global App_Version
    global lowest_date
    global latest_date

    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_CrestApp_Review.xlsx')

    # Calculate review counts
    positive_reviews_count = len(df[df['Ratings'].isin([4, 5])])
    negative_reviews_count = len(df[df['Ratings'].isin([1, 2, 3])])

    # Calculate summary statistics
    app_name = "Android Crest App"  # Replace with your app name
    app_rating = App_Rating + '*'
    app_version = App_Version
    review_data = str(lowest_date) + ' to ' + str(latest_date)

    # Define app details data
    app_details_data = {
        'App Name': [app_name],
        'App Rating': [app_rating],
        'Latest App Version': [app_version],  # Keep the app version unchanged
        'Review Date': [review_data]
    }

    # Create the app details DataFrame
    app_details_df = pd.DataFrame(app_details_data)

    # Create a DataFrame for the review counts
    posneg_review_counts_data = {
        'Category': ['Positive Reviews (4 & 5)', 'Negative Reviews (<=3)'],
        'Count': [positive_reviews_count, negative_reviews_count]
    }
    posneg_review_counts_df = pd.DataFrame(posneg_review_counts_data)

    # Create a DataFrame for the negative review analysis
    negative_reason_counts = Negative_Review_Analysis()
    negative_review_counts_df = pd.DataFrame({'Summary (-ve reviews)': negative_reason_counts.index, 'Count': negative_reason_counts.values})

    # Count the occurrences of 'Positive Feedback' in negative review analysis
    positive_feedback_count = negative_reason_counts.get('Positive Feedback', 0)

    # Ensure 'Count' column can hold string values
    negative_review_counts_df['Count'] = negative_review_counts_df['Count'].astype(object)

    # Update the 'Latest App Version' column for the 'Positive Feedback' row with the feedback message
    if positive_feedback_count > 0:
        feedback_message = f"Reviews are good but ratings are <= 3"

        # Find the row for 'Positive Feedback' in negative review counts DataFrame
        positive_feedback_row_index = negative_review_counts_df[negative_review_counts_df['Summary (-ve reviews)'] == 'Positive Feedback'].index[0]

        # Update the 'Latest App Version' column of 'Positive Feedback' with the feedback message
        negative_review_counts_df.at[positive_feedback_row_index, ''] = feedback_message  # Set the feedback message

        # Append this message to the Latest App Version in the app details
        app_details_df.at[0, 'Latest App Version'] = app_version  # Keep the app version unchanged
        #app_details_df.at[0, 'Latest App Version'] += f" (See feedback in negative reviews)"  # Optional: Add a note

    # Load the workbook
    wb = load_workbook('Android_CrestApp_Review.xlsx')

    # Check if the 'Summary' sheet exists
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        # Append app details and review counts to 'Summary' sheet
        for r in dataframe_to_rows(app_details_df, index=False):
            ws.append(r)
        for row in dataframe_to_rows(posneg_review_counts_df, index=False):
            ws.append(row)
        ws.append([])  # Add an empty row
        for row in dataframe_to_rows(negative_review_counts_df, index=False):
            ws.append(row)

    else:
        # Create a new 'Summary' sheet and write app details and review counts
        ws = wb.create_sheet(title='Summary')
        for r in dataframe_to_rows(app_details_df, index=False):
            ws.append(r)
        ws.append([])  # Add an empty row
        for row in dataframe_to_rows(posneg_review_counts_df, index=False):
            ws.append(row)
        ws.append([])  # Add an empty row
        for row in dataframe_to_rows(negative_review_counts_df, index=False):
            ws.append(row)

    # Save the workbook
    wb.save('Android_CrestApp_Review.xlsx')
    print('Summary sheet has been updated with app details, review counts, and negative review analysis.')

def Summary_Enhance():
    # Load the workbook
    wb = load_workbook('Android_CrestApp_Review.xlsx')

    # Select the "Summary" sheet
    sheet = wb['Summary']

    # Make the 1st row from column A to D bold and set background color to blue
    for col in range(1, 5):  # Columns A to D
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B9D9EB", end_color="B9D9EB", fill_type="solid")
        #cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    # Apply center alignment to columns B, C, and D
    for col in range(2, 5):  # Columns B to D
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Make the 4th row from column A to D bold and set background color
    for col in range(1, 3):  # Columns A to D
        cell = sheet.cell(row=4, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Make the 8th row from column A to D bold and set background color to yellow
    for col in range(1, 3):  # Columns A to D
        cell = sheet.cell(row=6, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Make the 8th row from column A to D bold and set background color
    for col in range(1, 3):  # Columns A to D (1 to 4 inclusive)
        cell = sheet.cell(row=5, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="66CC66", end_color="66CC66", fill_type="solid")

    # Make the 8th row from column A to D bold and set background color
    for col in range(1, 4):  # Columns A to D
        cell = sheet.cell(row=8, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Save the workbook (same file)
    wb.save('Android_CrestApp_Review.xlsx')
    print("Summary Sheet's Last changes has been updated")

def close_excel_file(file_path):
    """
    Close the Excel file to release any locks on it.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbook.close()
        print("Closed the Excel file.")
    except Exception as e:
        print(f"Error closing the Excel file: {e}")


if __name__ == '__main__':
    CrestApp_WebScraping()
    time.sleep(3)
    All_HtmlFiles_Csv()
    time.sleep(2)

    # IMPORTANT: Call App_Rating_Finder BEFORE Remove_Html()
    App_Rating_Finder()

    Csv_Date_Rating_Filter_7days()
    #Csv_Date_Rating_Filter_Alldate()
    CSV_Remove_Duplicates()
    Csv_2_Xlsx()
    SheetNameChange()
    Positive_Reviews()
    Negative_Reviews()
    Sorting_Xlsx('Positive_Reviews')
    Sorting_Xlsx('Negative_Reviews')
    Sorting_Xlsx('CrestApp_Review_Data')
    Negative_Review_Analysis()
    App_Version_Finder()
    Review_Date_Calculator()
    Summary()
    Sorting_Summary_Xlsx('Summary')
    time.sleep(3)
    Summary_Enhance()
    time.sleep(3)
    close_excel_file('Android_CrestApp_Review.xlsx')

    # Remove HTML files LAST, after all processing is done
    Remove_Html()

    # Specify file names and remove CSV files
    file1 = 'Crest_Data.csv'
    file2 = 'CrestAppGCWeb.csv'

    # Remove the files after execution is completed
    try:
        os.remove(file1)
        print(f"{file1} has been deleted successfully.")
    except FileNotFoundError:
        print(f"{file1} not found.")

    try:
        os.remove(file2)
        print(f"{file2} has been deleted successfully.")
    except FileNotFoundError:
        print(f"{file2} not found.")